"""Write a validated CompanyForecast into its OpenStocks workbook.

The template is authoritative. We open `challenge/templates/<file>.xlsx`, write
only the three yellow input cells, and save to `submission/`. Sheet name, metric
labels, units and the period header are never touched — the structural check
fails the submission if any of them move.

Cell layout, verified against all four supplied templates:

    row 6   Metric | Units | <Period>      header
    row 7   metric 1                        value goes in C7
    row 8   metric 2                        C8
    row 9   metric 3                        C9

Row order follows companies.json, which is also the order `submitted_specs()`
returns, so the two cannot drift apart silently.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from forecast.metrics import submitted_specs
from forecast.schema import CompanyForecast, Unit

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = REPO_ROOT / "challenge" / "templates"
SUBMISSION_DIR = REPO_ROOT / "submission"

SHEET = "Summary"
FIRST_METRIC_ROW = 7
VALUE_COLUMN = "C"
LABEL_COLUMN = "A"
UNITS_COLUMN = "B"

#: Decimal places per unit. Percentages and EPS carry two; money is whole
#: millions, because no company reports net sales to a fraction of a million and
#: false precision invites a reader to trust the number more than it deserves.
ROUNDING: dict[Unit, int] = {
    Unit.USD_M: 0,
    Unit.GBP_M: 1,
    Unit.USD_PER_SHARE: 2,
    Unit.GBP_PENCE: 2,
    Unit.PERCENT: 1,
}


class TemplateMismatch(RuntimeError):
    """The template does not look how the writer expects."""


def write_workbook(forecast: CompanyForecast) -> Path:
    """Write one company's three forecasts into its workbook.

    Every label and unit is checked against the template before anything is
    written, so a renamed metric fails here rather than at upload time.
    """
    template = TEMPLATE_DIR / forecast.output_file
    if not template.is_file():
        raise TemplateMismatch(f"template missing: {template}")

    workbook = load_workbook(template)
    if SHEET not in workbook.sheetnames:
        raise TemplateMismatch(f"{forecast.output_file}: no '{SHEET}' sheet")
    sheet = workbook[SHEET]

    specs = submitted_specs(forecast.company)
    by_label = {m.label: m for m in forecast.metrics}

    for offset, spec in enumerate(specs):
        row = FIRST_METRIC_ROW + offset

        template_label = sheet[f"{LABEL_COLUMN}{row}"].value
        template_units = sheet[f"{UNITS_COLUMN}{row}"].value
        if template_label != spec.label:
            raise TemplateMismatch(
                f"{forecast.output_file} row {row}: template says "
                f"{template_label!r}, registry says {spec.label!r}"
            )
        if template_units != spec.units.value:
            raise TemplateMismatch(
                f"{forecast.output_file} row {row}: template units "
                f"{template_units!r} != {spec.units.value!r}"
            )

        metric = by_label.get(spec.label)
        if metric is None:
            raise TemplateMismatch(
                f"{forecast.output_file}: no forecast for {spec.label!r}"
            )

        # Numbers only — no formulas, currency symbols or percent signs.
        sheet[f"{VALUE_COLUMN}{row}"] = round(
            metric.value, ROUNDING.get(metric.units, 2)
        )

    SUBMISSION_DIR.mkdir(parents=True, exist_ok=True)
    out = SUBMISSION_DIR / forecast.output_file
    workbook.save(out)
    return out
